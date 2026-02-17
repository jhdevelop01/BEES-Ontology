"""
보고서 생성 서비스
에너지, 장비 정비, 쾌적도, 알람 보고서 데이터 생성 로직.
"""

import csv
import io
import json
import logging
import os
import time
import uuid
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

from app.services import influxdb_service, mqtt_service, postgres_service

logger = logging.getLogger("server-a.reports")

# 보고서 프리셋 정의
REPORT_PRESETS = [
    {
        "id": "energy_monthly",
        "name": "에너지 월간 보고서",
        "description": "월간 에너지 소비량, 시스템별 비율, EUI 지표 포함",
    },
    {
        "id": "maintenance_status",
        "name": "장비 정비 현황",
        "description": "전체 작업 주문 현황, 완료율, 담당자별 실적",
    },
    {
        "id": "comfort",
        "name": "쾌적도 보고서",
        "description": "층별 온도/습도 분포, 쾌적 범위 이탈 현황",
    },
    {
        "id": "alarm_summary",
        "name": "알람 요약 보고서",
        "description": "기간별 알람 발생 건수, 심각도별 분류, 장비별 빈도",
    },
]

# 생성된 보고서 캐시 (메모리)
_report_cache: dict[str, dict[str, Any]] = {}


def get_presets() -> list[dict[str, Any]]:
    """프리셋 목록 반환."""
    return REPORT_PRESETS


async def generate_report(
    preset_id: str | None,
    custom: dict[str, Any] | None,
    period_start: str,
    period_end: str,
    fmt: str = "json",
) -> dict[str, Any]:
    """
    보고서 데이터 생성.
    CSV/JSON 형식으로 반환.
    """
    report_id = str(uuid.uuid4())[:8]
    generated_at = datetime.utcnow().isoformat()

    if preset_id == "energy_monthly":
        data = await _generate_energy_report(period_start, period_end)
    elif preset_id == "maintenance_status":
        data = await _generate_maintenance_report(period_start, period_end)
    elif preset_id == "comfort":
        data = await _generate_comfort_report(period_start, period_end)
    elif preset_id == "alarm_summary":
        data = await _generate_alarm_report(period_start, period_end)
    elif custom:
        data = await _generate_custom_report(custom, period_start, period_end)
    else:
        data = {"error": "preset_id 또는 custom 필요"}

    rows = data.get("rows", [])
    title = data.get("title", "보고서")
    summary = {k: v for k, v in data.items() if k not in ("rows", "title")}

    # 포맷별 변환
    csv_content = None
    binary_content = None
    content_type = None
    file_ext = fmt

    if fmt == "csv" and isinstance(rows, list):
        csv_content = _to_csv(rows)
    elif fmt == "xlsx" and isinstance(rows, list):
        binary_content = _to_xlsx(title, rows, summary)
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif fmt == "pdf" and isinstance(rows, list):
        binary_content = _to_pdf(title, rows, summary)
        content_type = "application/pdf"

    report = {
        "report_id": report_id,
        "preset_id": preset_id,
        "period": {"start": period_start, "end": period_end},
        "format": fmt,
        "status": "ready",
        "generated_at": generated_at,
        "data": data if fmt == "json" else None,
        "csv_content": csv_content,
        "binary_content": binary_content,
        "content_type": content_type,
        "file_ext": file_ext,
    }

    _report_cache[report_id] = report
    logger.info("보고서 생성 완료: %s (preset=%s, format=%s)", report_id, preset_id, fmt)
    return report


def get_report(report_id: str) -> dict[str, Any] | None:
    """캐시에서 보고서 조회."""
    return _report_cache.get(report_id)


def get_report_history(page: int = 1, limit: int = 20) -> dict[str, Any]:
    """생성된 보고서 이력."""
    all_reports = sorted(
        _report_cache.values(),
        key=lambda r: r.get("generated_at", ""),
        reverse=True,
    )
    start = (page - 1) * limit
    end = start + limit
    items = [
        {k: v for k, v in r.items() if k not in ("data", "csv_content", "binary_content")}
        for r in all_reports[start:end]
    ]
    return {
        "items": items,
        "total": len(all_reports),
        "page": page,
        "pages": (len(all_reports) + limit - 1) // limit,
    }


# ── 내부 보고서 생성 함수 ──────────────────────────────────────────────────

async def _generate_energy_report(start: str, end: str) -> dict[str, Any]:
    """에너지 보고서 데이터."""
    point_cache = mqtt_service.get_point_cache()
    power_points = {
        pid: data for pid, data in point_cache.items()
        if any(kw in pid.lower() for kw in ["power", "kw", "watt"])
    }
    total_kw = sum(d.get("value", 0) for d in power_points.values() if isinstance(d.get("value"), (int, float)))
    return {
        "title": "에너지 월간 보고서",
        "total_kw": round(total_kw, 2),
        "power_point_count": len(power_points),
        "rows": [
            {"point_id": pid, "value_kw": round(d.get("value", 0), 2), "unit": d.get("unit", "")}
            for pid, d in power_points.items()
        ],
    }


async def _generate_maintenance_report(start: str, end: str) -> dict[str, Any]:
    """장비 정비 보고서 데이터."""
    pool = postgres_service.get_pool()
    if not pool:
        return {"title": "장비 정비 현황", "error": "PostgreSQL 미연결", "rows": []}

    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT wo.id, wo.title, wo.status, wo.priority, wo.created_at, wo.completed_at,
                   em.ontology_id, u.name AS assigned_to_name
            FROM work_orders wo
            LEFT JOIN equipment_metadata em ON wo.equipment_id = em.id
            LEFT JOIN users u ON wo.assigned_to = u.id
            ORDER BY wo.created_at DESC LIMIT 100
            """
        )

    items = []
    for r in rows:
        items.append({
            "id": r["id"],
            "title": r["title"],
            "status": r["status"],
            "priority": r["priority"],
            "equipment": r["ontology_id"],
            "assigned_to": r["assigned_to_name"],
            "created_at": r["created_at"].isoformat() if r["created_at"] else "",
            "completed_at": r["completed_at"].isoformat() if r["completed_at"] else "",
        })

    completed = sum(1 for i in items if i["status"] == "completed")
    return {
        "title": "장비 정비 현황",
        "total": len(items),
        "completed": completed,
        "completion_rate": round(completed / max(len(items), 1) * 100, 1),
        "rows": items,
    }


async def _generate_comfort_report(start: str, end: str) -> dict[str, Any]:
    """쾌적도 보고서 데이터."""
    point_cache = mqtt_service.get_point_cache()
    temp_data = []
    humidity_data = []
    for pid, data in point_cache.items():
        value = data.get("value")
        if not isinstance(value, (int, float)):
            continue
        if "temp" in pid.lower():
            temp_data.append({"point_id": pid, "value": value, "unit": "°C"})
        elif "humid" in pid.lower() or "rh" in pid.lower():
            humidity_data.append({"point_id": pid, "value": value, "unit": "%"})

    avg_temp = round(sum(d["value"] for d in temp_data) / max(len(temp_data), 1), 1)
    avg_humidity = round(sum(d["value"] for d in humidity_data) / max(len(humidity_data), 1), 1)

    return {
        "title": "쾌적도 보고서",
        "avg_temperature": avg_temp,
        "avg_humidity": avg_humidity,
        "temp_sensor_count": len(temp_data),
        "humidity_sensor_count": len(humidity_data),
        "rows": temp_data + humidity_data,
    }


async def _generate_alarm_report(start: str, end: str) -> dict[str, Any]:
    """알람 요약 보고서 데이터."""
    alarms = mqtt_service.get_alarm_cache()
    by_severity: dict[str, int] = {}
    by_equipment: dict[str, int] = {}
    for a in alarms:
        sev = a.get("severity", "unknown")
        by_severity[sev] = by_severity.get(sev, 0) + 1
        equip = a.get("equipment", "unknown")
        by_equipment[equip] = by_equipment.get(equip, 0) + 1

    return {
        "title": "알람 요약 보고서",
        "total_alarms": len(alarms),
        "by_severity": by_severity,
        "by_equipment": dict(sorted(by_equipment.items(), key=lambda x: -x[1])[:20]),
        "rows": [
            {"severity": a.get("severity"), "equipment": a.get("equipment"),
             "message": a.get("message", ""), "ts": a.get("ts")}
            for a in alarms
        ],
    }


async def _generate_custom_report(custom: dict[str, Any], start: str, end: str) -> dict[str, Any]:
    """사용자 정의 보고서."""
    return {
        "title": "사용자 정의 보고서",
        "custom_config": custom,
        "rows": [],
    }


def _to_csv(rows: list[dict[str, Any]]) -> str:
    """딕셔너리 리스트를 CSV 문자열로 변환."""
    if not rows:
        return ""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=rows[0].keys())
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue()


def _to_xlsx(title: str, rows: list[dict[str, Any]], summary: dict[str, Any]) -> bytes:
    """딕셔너리 리스트를 Excel(.xlsx) 바이너리로 변환."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: 요약 ──
    ws_summary = wb.active
    ws_summary.title = "요약"

    header_font = Font(bold=True, size=14)
    label_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_text = Font(bold=True, color="FFFFFF", size=11)

    ws_summary["A1"] = f"BEES 디지털 트윈 — {title}"
    ws_summary["A1"].font = header_font
    ws_summary.merge_cells("A1:D1")

    ws_summary["A3"] = "생성일시"
    ws_summary["A3"].font = label_font
    ws_summary["B3"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")

    row_idx = 5
    for key, value in summary.items():
        if isinstance(value, (dict, list)):
            ws_summary.cell(row=row_idx, column=1, value=key).font = label_font
            ws_summary.cell(row=row_idx, column=2, value=json.dumps(value, ensure_ascii=False))
        else:
            ws_summary.cell(row=row_idx, column=1, value=key).font = label_font
            ws_summary.cell(row=row_idx, column=2, value=value)
        row_idx += 1

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 50

    # ── Sheet 2: 상세 데이터 ──
    if rows:
        ws_data = wb.create_sheet("상세 데이터")
        headers = list(rows[0].keys())

        for col_idx, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.font = header_text
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                ws_data.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

        # 컬럼 너비 자동 조정
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row_data in rows[:50]:
                val = str(row_data.get(header, ""))
                max_len = max(max_len, len(val))
            ws_data.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _to_pdf(title: str, rows: list[dict[str, Any]], summary: dict[str, Any]) -> bytes:
    """딕셔너리 리스트를 PDF 바이너리로 변환."""
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # 한글 폰트 로딩 (NanumGothic → Helvetica 폴백)
    _FONT_PATHS = [
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",         # Debian/Docker
        "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf",
        "/System/Library/Fonts/AppleSDGothicNeo.ttc",              # macOS
    ]
    font_family = "Helvetica"
    for fpath in _FONT_PATHS:
        if os.path.exists(fpath):
            try:
                pdf.add_font("nanum", "", fpath)
                bold_path = fpath.replace("NanumGothic.ttf", "NanumGothicBold.ttf")
                if os.path.exists(bold_path):
                    pdf.add_font("nanum", "B", bold_path)
                else:
                    pdf.add_font("nanum", "B", fpath)
                font_family = "nanum"
                break
            except Exception:
                pass

    # ── 헤더 ──
    pdf.set_font(font_family, "B", 16)
    pdf.set_fill_color(31, 78, 121)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 12, f"BEES 디지털 트윈 — {title}", new_x="LMARGIN", new_y="NEXT", fill=True, align="C")
    pdf.ln(3)

    pdf.set_text_color(0, 0, 0)
    pdf.set_font(font_family, "", 9)
    pdf.cell(0, 6, f"생성일시: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # ── 요약 섹션 ──
    if summary:
        pdf.set_font(font_family, "B", 11)
        pdf.cell(0, 8, "요약", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font(font_family, "", 9)
        for key, value in summary.items():
            if isinstance(value, (dict, list)):
                text = json.dumps(value, ensure_ascii=False)
            else:
                text = str(value)
            pdf.cell(60, 6, str(key), border=1)
            pdf.cell(0, 6, text[:100], border=1, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(5)

    # ── 데이터 테이블 ──
    if rows:
        headers = list(rows[0].keys())
        num_cols = len(headers)
        page_width = pdf.w - pdf.l_margin - pdf.r_margin
        col_width = page_width / max(num_cols, 1)

        pdf.set_font(font_family, "B", 11)
        pdf.cell(0, 8, "상세 데이터", new_x="LMARGIN", new_y="NEXT")

        # 테이블 헤더
        pdf.set_font(font_family, "B", 8)
        pdf.set_fill_color(31, 78, 121)
        pdf.set_text_color(255, 255, 255)
        for header in headers:
            pdf.cell(col_width, 7, str(header)[:15], border=1, fill=True, align="C")
        pdf.ln()

        # 테이블 행
        pdf.set_font(font_family, "", 7)
        pdf.set_text_color(0, 0, 0)
        for i, row_data in enumerate(rows):
            if i % 2 == 0:
                pdf.set_fill_color(240, 240, 240)
            else:
                pdf.set_fill_color(255, 255, 255)
            for header in headers:
                val = str(row_data.get(header, ""))[:20]
                pdf.cell(col_width, 6, val, border=1, fill=True, align="C")
            pdf.ln()

    # ── 푸터 ──
    pdf.ln(10)
    pdf.set_font(font_family, "", 7)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 5, "삼성물산 GEC B동 디지털 트윈 플랫폼 — BEES", align="C")

    return bytes(pdf.output())
